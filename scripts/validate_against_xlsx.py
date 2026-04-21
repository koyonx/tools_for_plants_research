#!/usr/bin/env python3
"""Compare basic_measurement output against the legacy `measure_results.xlsx`.

Reads the spreadsheet, looks up each filename in Supabase by
`original_filename`, kicks off the backend's `POST /images/{id}/analyze`
basic-measurement endpoint, then computes per-row deltas and overall
stats.  Writes a Markdown report to `outputs/validation_report.md`.

Authentication
--------------
GoTrue magic-link accounts have no password, so password-grant only
works for accounts the operator explicitly provisioned with one.  Three
auth modes are supported, in priority order:

1. `--access-token <jwt>` — paste a fresh JWT from the running browser
   session (DevTools → Application → Local Storage → `sb-...-auth-token`
   → `access_token`).  No password ever touches the script.
2. `--user-email <addr>` + password prompt — only useful for accounts
   that have a password (created via Studio's auth UI or seeded
   manually).  Wrong password 4xx's loudly.
3. `--service-role` — reads `SERVICE_ROLE_KEY` from env and bypasses
   RLS entirely.  Suitable for CI / unattended runs; **never use this
   in shared production environments**.

Usage
-----
    # 1) Start the stack and upload the images that appear in the xlsx.
    make up

    # 2) Run the validator.  `make validate` sources .env automatically
    #    so the script sees ANON_KEY / SUPABASE_PUBLIC_URL etc.
    VALIDATE_EMAIL=you@example.com make validate
    # or, with a paste-in token:
    set -a; . .env; set +a
    python scripts/validate_against_xlsx.py \\
        --xlsx ../measure_results.xlsx \\
        --reference-um 100 \\
        --access-token <paste from devtools>

Outputs land at `outputs/validation_report.md` + `.json` *relative to
the cwd you ran the script from* (typically the repo root via `make`).
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
import time
from pathlib import Path
from typing import Any

import httpx
from openpyxl import load_workbook

# Heuristic mapping from the xlsx `メモ` column to backend metric keys.
# Override with `--metric-map '{"厚さ":"leaf_mean_thickness_um"}'` to
# tweak per-experiment without editing the script.
DEFAULT_METRIC_MAP = {
    "厚さ": "leaf_mean_thickness_um",  # full leaf thickness
    "葉肉の厚さ": "leaf_median_thickness_um",  # mesophyll-only ~ central
}
AREA_METRIC = "leaf_area_um2"
AREA_NOTE = "維管束面積"  # not directly available in basic_measurement

VALUE_RE = re.compile(r"^([\-+]?\d+(?:\.\d+)?)\s*(?:µm|um|μm)?\s*(?:\^?[²2])?$")


def _parse_value(raw: str | float | int | None) -> float | None:
    if raw is None:
        return None
    if isinstance(raw, int | float):
        return float(raw)
    text = str(raw).strip().replace(",", "")
    # Normalise unicode minus / space-less µm units before regexing.
    text = text.replace("\u2212", "-").replace("\u00a0", " ")
    m = VALUE_RE.match(text.replace(" ", ""))
    if m:
        try:
            return float(m.group(1))
        except ValueError:
            return None
    return None


def _read_groundtruth(xlsx_path: Path) -> list[dict[str, Any]]:
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb["All_Data"] if "All_Data" in wb.sheetnames else wb.active
    rows: list[dict[str, Any]] = []
    headers: list[str] | None = None
    for raw in ws.iter_rows(values_only=True):
        if headers is None:
            headers = [str(h) if h is not None else "" for h in raw]
            continue
        record = dict(zip(headers, raw, strict=False))
        rows.append(record)
    return rows


def _looks_like_jwt(token: str) -> bool:
    return token.count(".") == 2 and len(token) > 100


def _signin_password(supabase_url: str, anon_key: str, email: str, password: str) -> str:
    try:
        resp = httpx.post(
            f"{supabase_url}/auth/v1/token?grant_type=password",
            headers={"apikey": anon_key, "Content-Type": "application/json"},
            json={"email": email, "password": password},
            timeout=30,
        )
        resp.raise_for_status()
        body = resp.json()
    except (httpx.HTTPError, ValueError) as e:
        raise SystemExit(
            "password sign-in failed; magic-link accounts have no password — "
            "use --access-token instead.  underlying error: " + str(e)
        ) from e
    token = body.get("access_token") if isinstance(body, dict) else None
    if not token:
        raise SystemExit("GoTrue did not return an access_token; aborting.")
    return str(token)


def _find_image(
    supabase_url: str, anon_key: str, jwt: str, filename: str
) -> dict[str, Any] | None:
    """Look up an image by `original_filename`.

    Raises `RuntimeError` when more than one row matches — a single
    `original_filename` can be reused across uploaders, especially when
    running with `--service-role`, and silently picking the first match
    would skew the validation report.
    """
    try:
        resp = httpx.get(
            f"{supabase_url}/rest/v1/images",
            params={"original_filename": f"eq.{filename}", "select": "*"},
            headers={"apikey": anon_key, "Authorization": f"Bearer {jwt}"},
            timeout=30,
        )
        resp.raise_for_status()
        rows = resp.json()
    except (httpx.HTTPError, ValueError) as e:
        raise RuntimeError(f"image lookup failed: {e}") from e
    if not isinstance(rows, list) or not rows:
        return None
    if len(rows) > 1:
        ids = ", ".join(str(r.get("id"))[:8] for r in rows)
        raise RuntimeError(
            f"{len(rows)} images share the filename {filename!r} (ids: {ids}); "
            "constrain by uploader/visibility before re-running."
        )
    first = rows[0]
    return first if isinstance(first, dict) else None


def _run_basic_measurement(
    backend_url: str, jwt: str, image_id: str, reference_um: float
) -> dict[str, Any]:
    try:
        resp = httpx.post(
            f"{backend_url}/images/{image_id}/analyze",
            headers={"Authorization": f"Bearer {jwt}", "Content-Type": "application/json"},
            json={"reference_um": reference_um},
            timeout=180,
        )
        resp.raise_for_status()
        body = resp.json()
    except (httpx.HTTPError, ValueError) as e:
        raise RuntimeError(f"basic_measurement failed: {e}") from e
    if not isinstance(body, dict):
        raise RuntimeError(f"basic_measurement returned non-dict: {body!r}")
    return body


def _format_delta(measured: float | None, expected: float | None) -> str:
    if measured is None or expected is None:
        return "—"
    delta = measured - expected
    pct = 100.0 * delta / expected if expected else 0.0
    return f"{delta:+.2f} ({pct:+.1f}%)"


def main() -> int:
    parser = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawTextHelpFormatter
    )
    parser.add_argument("--xlsx", required=True, type=Path, help="path to measure_results.xlsx")
    parser.add_argument(
        "--reference-um", type=float, default=100.0, help="scale-bar length in µm"
    )
    parser.add_argument(
        "--out", type=Path, default=Path("outputs/validation_report.md"),
        help="report destination, resolved against cwd",
    )
    parser.add_argument(
        "--supabase-url",
        default=os.environ.get("SUPABASE_PUBLIC_URL") or "http://localhost:8000",
    )
    parser.add_argument(
        "--anon-key",
        default=os.environ.get("ANON_KEY") or os.environ.get("NEXT_PUBLIC_SUPABASE_ANON_KEY") or "",
    )
    parser.add_argument(
        "--backend-url",
        default=os.environ.get("NEXT_PUBLIC_BACKEND_URL") or "http://localhost:8001",
    )
    parser.add_argument(
        "--metric-map",
        default=None,
        help="JSON object overriding DEFAULT_METRIC_MAP, e.g. "
        '\'{"厚さ":"leaf_mean_thickness_um"}\'',
    )
    auth_group = parser.add_mutually_exclusive_group(required=True)
    auth_group.add_argument(
        "--access-token",
        help="Paste a fresh user JWT (e.g. from browser DevTools).  No password needed.",
    )
    auth_group.add_argument(
        "--user-email",
        help="Email for password grant (only works for accounts WITH a password)",
    )
    auth_group.add_argument(
        "--service-role",
        action="store_true",
        help="Use SERVICE_ROLE_KEY env var (bypasses RLS — never in shared prod)",
    )
    args = parser.parse_args()

    if not args.anon_key:
        parser.error(
            "anon key not provided — set ANON_KEY (or run via `make validate`, "
            "which sources .env automatically), or pass --anon-key."
        )

    metric_map = dict(DEFAULT_METRIC_MAP)
    if args.metric_map:
        try:
            metric_map.update(json.loads(args.metric_map))
        except json.JSONDecodeError as e:
            parser.error(f"--metric-map is not valid JSON: {e}")

    if args.access_token:
        jwt = args.access_token
        if not _looks_like_jwt(jwt):
            parser.error("--access-token does not look like a JWT (need three dot-separated parts)")
    elif args.service_role:
        jwt = os.environ.get("SERVICE_ROLE_KEY", "")
        if not jwt:
            parser.error("--service-role requires SERVICE_ROLE_KEY in env")
        if not _looks_like_jwt(jwt):
            parser.error("SERVICE_ROLE_KEY does not look like a JWT")
        print("WARNING: --service-role bypasses RLS.  Do not use in shared prod.", file=sys.stderr)
    else:
        import getpass

        password = getpass.getpass(f"password for {args.user_email}: ")
        jwt = _signin_password(args.supabase_url, args.anon_key, args.user_email, password)

    rows = _read_groundtruth(args.xlsx)
    print(f"loaded {len(rows)} ground-truth rows from {args.xlsx}", file=sys.stderr)
    print(f"metric mapping: {metric_map}", file=sys.stderr)

    grouped: dict[str, list[dict[str, Any]]] = {}
    for r in rows:
        fname = r.get("ファイル名")
        if not fname:
            continue
        grouped.setdefault(str(fname), []).append(r)

    output_lines: list[str] = []
    output_lines.append("# Validation report\n")
    output_lines.append(f"- xlsx: `{args.xlsx}`")
    output_lines.append(f"- reference_um: {args.reference_um}")
    output_lines.append(f"- generated: {time.strftime('%Y-%m-%d %H:%M:%S')}\n")

    summary_rows: list[tuple[str, str, float | None, float | None, str]] = []

    for filename, gt_rows in grouped.items():
        output_lines.append(f"## {filename}\n")
        try:
            image = _find_image(args.supabase_url, args.anon_key, jwt, filename)
        except RuntimeError as e:
            output_lines.append(f"- ERROR: {e}\n")
            continue
        if image is None:
            output_lines.append(
                f"- ERROR: image `{filename}` not found in `images` table; upload it first\n"
            )
            continue

        try:
            analysis = _run_basic_measurement(
                args.backend_url, jwt, image["id"], args.reference_um
            )
        except RuntimeError as e:
            output_lines.append(f"- ERROR: {e}\n")
            continue

        result = analysis.get("result") or {}
        measurement = result.get("measurement") or {}
        scale = result.get("scale") or {}
        output_lines.append(f"- µm/px detected: `{scale.get('um_per_px', 'n/a')}`")
        output_lines.append(f"- analysis_id: `{analysis.get('id')}`\n")

        output_lines.append("| メモ | 期待値 (µm/µm²) | 実測 (µm/µm²) | Δ (差) |")
        output_lines.append("| --- | ---: | ---: | ---: |")
        for gt in gt_rows:
            note = str(gt.get("メモ", "")).strip()
            mtype = str(gt.get("測定タイプ", "")).strip()
            expected = _parse_value(gt.get("測定値"))
            measured: float | None = None
            if note in metric_map:
                measured = measurement.get(metric_map[note])
            elif note == AREA_NOTE and mtype == "Area":
                # 維管束面積 isn't measured by basic_measurement; compare
                # against the SegFormer xylem polygon area instead (out
                # of scope for this script).
                measured = None
            elif mtype == "Area":
                measured = measurement.get(AREA_METRIC)
            output_lines.append(
                f"| {note} | {expected if expected is not None else '—'} "
                f"| {f'{measured:.2f}' if measured is not None else '—'} "
                f"| {_format_delta(measured, expected)} |"
            )
            summary_rows.append(
                (filename, note, measured, expected, _format_delta(measured, expected))
            )
        output_lines.append("")

    # Aggregate stats
    output_lines.append("## Aggregate\n")
    parsed = [
        (m, e) for _, _, m, e, _ in summary_rows if m is not None and e is not None and e != 0
    ]
    if not parsed:
        output_lines.append("- no rows with both expected + measured values\n")
    else:
        rel_errors = [abs(m - e) / e for m, e in parsed]
        mae = sum(abs(m - e) for m, e in parsed) / len(parsed)
        output_lines.append(f"- compared rows: {len(parsed)}")
        output_lines.append(f"- mean absolute error: {mae:.2f}")
        output_lines.append(f"- mean relative error: {100 * sum(rel_errors) / len(parsed):.1f} %")
        output_lines.append(f"- max relative error: {100 * max(rel_errors):.1f} %\n")

    args.out.parent.mkdir(parents=True, exist_ok=True)
    args.out.write_text("\n".join(output_lines), encoding="utf-8")
    args.out.with_suffix(".json").write_text(
        json.dumps(
            [
                {"filename": f, "note": n, "measured": m, "expected": e}
                for f, n, m, e, _ in summary_rows
            ],
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    print(f"wrote {args.out}", file=sys.stderr)
    print(f"wrote {args.out.with_suffix('.json')}", file=sys.stderr)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
