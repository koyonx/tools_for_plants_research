"""Tests for the LI-COR gas-exchange parser.

Synthetic files are built in memory for each case so CI doesn't need
real field data.  The fixtures cover the two common variants we've
seen in the wild:

- LI-6400 style: plain tab-delimited text with a small header block,
  column names like ``Obs``, ``Photo``, ``Cond``, ``Ci``, ``PARi``.
- LI-6800 style: .xlsx with ``HEADER`` / ``DATA`` markers, a unit
  row below the column headers, column names like ``A``, ``gsw``,
  ``Ci``, ``Qin``.

Plus an adversarial CSV with operator-renamed columns and stray blank
rows, and a negative case confirming the parser reports cleanly when
no header row can be identified.
"""

from __future__ import annotations

import io
import json
from datetime import datetime

import pytest
from openpyxl import Workbook

from app.pipeline.licor_parse import (
    LicorParseError,
    parse_delimited,
    parse_file,
    parse_xlsx,
)


def _li6400_tsv(num_points: int = 5) -> bytes:
    """Minimal LI-6400 tab-delimited export."""
    lines = [
        "\"OPEN 6.2.4 (LI-COR Inc., Lincoln, Nebraska, USA)\"",
        "\"LI-6400 Portable Photosynthesis System\"",
        "",
        "FileDate\t2025-07-14 09:32:11",
        "Operator\tkoyon",
        "",
        "Obs\tHHMMSS\tPhoto\tCond\tCi\tTleaf\tPARi\tCO2R\tVpdL\tTrmmol",
    ]
    for i in range(num_points):
        lines.append(
            f"{i + 1}\t09:{32 + i:02d}:00\t{15.0 - i:.3f}\t{0.25 + 0.01 * i:.4f}"
            f"\t{260.0 + 5 * i:.1f}\t{25.0 + 0.1 * i:.2f}\t1500.0\t400.0\t1.15\t3.2"
        )
    return "\n".join(lines).encode("utf-8")


def _li6800_xlsx(num_points: int = 4) -> bytes:
    """LI-6800 Excel export: HEADER block, column names, unit row, data."""
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Measurements"
    ws.append(["LI-6800 Portable Photosynthesis System"])
    ws.append(["HEADER"])
    ws.append(["Instrument", "LI-6800"])
    ws.append(["Firmware", "2.1.03"])
    ws.append(["Started", "2025-08-02T10:15:00"])
    ws.append([])
    ws.append(["obs", "time", "A", "gsw", "Ci", "Qin", "Tleaf", "VPDleaf", "CO2_r", "E"])
    ws.append(["-", "hh:mm:ss", "µmol m-2 s-1", "mol m-2 s-1", "µmol mol-1",
               "µmol m-2 s-1", "°C", "kPa", "µmol mol-1", "mmol m-2 s-1"])
    for i in range(num_points):
        ws.append([
            i + 1,
            # openpyxl rejects tz-aware datetimes when writing; the
            # parser promotes naive datetimes to UTC at read time.
            datetime(2025, 8, 2, 10, 15 + i, 0),
            12.0 + 0.5 * i,
            0.18 + 0.005 * i,
            240.0 + 10 * i,
            1500.0,
            26.5,
            1.35,
            410.0,
            2.8,
        ])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _custom_csv() -> bytes:
    """Operator-edited export: unusual order, different delimiter, extra custom columns."""
    return (
        b"# Experiment: C3 vs C4 drought\n"
        b"# Date: 2025-09-01\n"
        b"\n"
        b"obs_num,time,Photo (umol/m2/s),Cond,Ci,Custom_flag\n"
        b"1,2025-09-01 08:00:00,18.3,0.28,255,ok\n"
        b"2,2025-09-01 08:01:00,17.9,0.27,262,ok\n"
        b"3,2025-09-01 08:02:00,17.1,0.26,268,flagged\n"
    )


def test_li6400_tsv_parses_with_header_and_points() -> None:
    data = _li6400_tsv(num_points=5)
    session = parse_delimited(data, file_name="plant07.xls")
    assert session.instrument == "li_6400"
    assert session.source_format == "tsv"
    assert len(session.points) == 5
    first = session.points[0]
    assert first.photo_a == pytest.approx(15.0)
    assert first.cond_gsw == pytest.approx(0.25)
    assert first.ci_ppm == pytest.approx(260.0)
    assert first.leaf_temp_c == pytest.approx(25.0)
    assert first.par_umol == pytest.approx(1500.0)
    assert first.vpd_kpa == pytest.approx(1.15)
    assert first.transpiration == pytest.approx(3.2)
    # Header metadata is preserved verbatim (raw string, not coerced),
    # and the parser also infers a session-level captured_at from it.
    assert session.metadata.get("FileDate") == "2025-07-14 09:32:11"
    assert session.captured_at is not None
    assert "2025-07-14" in session.captured_at


def test_li6800_xlsx_skips_unit_row_and_infers_instrument() -> None:
    data = _li6800_xlsx(num_points=4)
    session = parse_xlsx(data, file_name="expt2.xlsx")
    assert session.instrument == "li_6800"
    assert session.source_format == "xlsx"
    assert len(session.points) == 4, (
        f"expected 4 points, got {len(session.points)} — unit row probably not skipped"
    )
    p0 = session.points[0]
    assert p0.photo_a == pytest.approx(12.0)
    assert p0.cond_gsw == pytest.approx(0.18)
    assert p0.par_umol == pytest.approx(1500.0)
    # Timestamp round-trip
    assert p0.recorded_at is not None
    assert "2025-08-02" in p0.recorded_at
    # Unit row markers must NOT leak into the body
    for point in session.points:
        assert isinstance(point.photo_a, float)


def test_custom_csv_with_edited_columns() -> None:
    data = _custom_csv()
    session = parse_delimited(data, file_name="custom.csv")
    assert session.source_format == "csv"
    assert len(session.points) == 3
    assert session.points[0].photo_a == pytest.approx(18.3)
    # Unknown columns round-trip through the raw blob so future PDE
    # analyses can still reach them.
    assert session.points[0].raw.get("Custom_flag") == "ok"
    assert session.points[2].raw.get("Custom_flag") == "flagged"


def test_missing_header_raises_clean_error() -> None:
    data = b"\n".join(
        [
            b"no columns here",
            b"just some free-form text",
            b"1,2,3",
            b"4,5,6",
        ]
    )
    with pytest.raises(LicorParseError, match="header row"):
        parse_delimited(data)


def test_parse_file_dispatches_on_magic_byte() -> None:
    """A bytes blob starting with PK (zip marker) should be treated as
    xlsx regardless of filename — matches reality where operators
    sometimes strip the extension."""
    data = _li6800_xlsx(num_points=2)
    session = parse_file(data, file_name="weird_name_no_extension")
    assert session.source_format == "xlsx"
    assert len(session.points) == 2


def test_legacy_xls_rejected_with_actionable_error() -> None:
    with pytest.raises(LicorParseError, match="legacy"):
        parse_file(b"irrelevant", file_name="old.xls")


def test_parsed_points_round_trip_through_strict_json() -> None:
    """The upload endpoint hands the parsed blob to Supabase as JSON;
    non-finite floats would be rejected server-side, so the parser
    must scrub NaN/Inf even when the source row contains them."""
    tsv_with_nan = (
        b"Obs\tPhoto\tCond\tCi\tPARi\n"
        b"1\tnan\t0.25\t260\t1500\n"
        b"2\t14.5\tinf\t255\t1500\n"
        b"3\t13.9\t0.23\tNA\t1500\n"
    )
    session = parse_delimited(tsv_with_nan)
    payload = [p.to_dict() for p in session.points]
    # None substitutes for NaN/Inf/NA — strict JSON accepts.
    s = json.dumps({"points": payload}, allow_nan=False)
    assert "NaN" not in s and "Infinity" not in s
    # Specifically check each rejected value became None
    assert session.points[0].photo_a is None
    assert session.points[1].cond_gsw is None
    assert session.points[2].ci_ppm is None


def test_blank_trailing_rows_do_not_break_header_detection() -> None:
    data = (
        b"some metadata\n"
        b"\n"
        b"Obs,Photo,Cond,Ci,PARi\n"
        b"1,10,0.2,250,1500\n"
        b"\n"
        b"\n"
        b"\n"
    )
    session = parse_delimited(data)
    assert len(session.points) == 1
    assert session.points[0].photo_a == pytest.approx(10.0)


def test_obs_and_timestamp_columns_are_extracted_and_not_in_raw() -> None:
    """`Obs` and `HHMMSS` map to the scalar fields, not to the raw blob,
    so the blob stays compact and downstream analytics aren't forced to
    de-duplicate the timestamp."""
    session = parse_delimited(_li6400_tsv(num_points=3))
    p = session.points[0]
    assert p.obs_index == 1
    assert "Obs" not in p.raw and "HHMMSS" not in p.raw


def test_multi_sheet_workbook_picks_highest_scoring_sheet() -> None:
    """A workbook where the first sheet is a summary (no LI-COR tokens)
    and the second is the raw log — parse the raw log, not the summary."""
    wb = Workbook()
    summary = wb.active
    assert summary is not None
    summary.title = "Summary"
    summary.append(["Report", "Generated 2025-10-01"])
    summary.append(["Mean photo", 14.3])
    data = wb.create_sheet("Raw")
    data.append(["HEADER"])
    data.append(["obs", "A", "gsw", "Ci", "Qin", "Tleaf"])
    for i in range(6):
        data.append([i + 1, 13.0 + 0.2 * i, 0.2, 250.0 + i, 1500.0, 25.5])
    buf = io.BytesIO()
    wb.save(buf)
    session = parse_xlsx(buf.getvalue())
    assert len(session.points) == 6
    assert session.instrument == "li_6800"


def test_duplicate_alias_columns_first_wins_loser_in_raw() -> None:
    """An operator-edited file with BOTH ``Photo`` AND ``A`` columns
    must keep the FIRST occurrence as the typed `photo_a` and route
    the duplicate's value into `raw` so nothing is silently lost.
    The session also gets a diagnostic note describing the collision.
    """
    data = (
        b"Obs,Photo,A,Cond,Ci,PARi\n"
        b"1,15.0,99.9,0.25,260,1500\n"
        b"2,14.5,99.5,0.24,265,1500\n"
    )
    session = parse_delimited(data)
    assert session.points[0].photo_a == pytest.approx(15.0)
    # Duplicate `A` value flows into raw under its original header.
    assert session.points[0].raw.get("A") == pytest.approx(99.9)
    assert session.notes is not None
    assert "duplicate alias" in session.notes
    assert "photo_a" in session.notes


def test_ambiguous_slash_date_returns_none_no_silent_misclassification() -> None:
    """A slash date where both leading fields are <= 12 is ambiguous
    (M/D/Y vs D/M/Y).  Refuse to guess — silently picking one
    convention would shift A-Ci measurements by months in operator
    files from non-en-US locales."""
    # 03/04/2025 could mean March 4 or April 3 — return None.
    data = (
        b"Obs,time,Photo,Cond,Ci,PARi\n"
        b"1,03/04/2025 10:00:00,15.0,0.25,260,1500\n"
    )
    session = parse_delimited(data)
    assert session.points[0].recorded_at is None


def test_unambiguous_slash_date_with_day_over_12_parses_as_dd_mm() -> None:
    # 25/04/2025 → can only mean Apr 25.  Parser must not silently
    # produce a different month.
    data = (
        b"Obs,time,Photo,Cond,Ci,PARi\n"
        b"1,25/04/2025 10:00:00,15.0,0.25,260,1500\n"
    )
    session = parse_delimited(data)
    assert session.points[0].recorded_at is not None
    assert "2025-04-25" in session.points[0].recorded_at


def test_unambiguous_slash_date_with_month_over_12_parses_as_mm_dd() -> None:
    # 04/25/2025 → can only mean Apr 25 with MM/DD ordering.
    data = (
        b"Obs,time,Photo,Cond,Ci,PARi\n"
        b"1,04/25/2025 10:00:00,15.0,0.25,260,1500\n"
    )
    session = parse_delimited(data)
    assert session.points[0].recorded_at is not None
    assert "2025-04-25" in session.points[0].recorded_at


def test_dotted_european_date_unambiguous() -> None:
    # 03.04.2025 (D.M.Y) is unambiguous — we accept it directly.
    data = (
        b"Obs,time,Photo,Cond,Ci,PARi\n"
        b"1,03.04.2025 10:00:00,15.0,0.25,260,1500\n"
    )
    session = parse_delimited(data)
    assert session.points[0].recorded_at is not None
    assert "2025-04-03" in session.points[0].recorded_at


def test_all_nan_body_emits_diagnostic_note() -> None:
    """A file whose body rows have header but every cell is empty / NaN
    should produce zero points AND a note explaining what happened —
    operator shouldn't see "0 points" with no context."""
    data = (
        b"Obs,Photo,Cond,Ci,PARi\n"
        b"1,NaN,NaN,NaN,NaN\n"
        b"2,,,,\n"
    )
    session = parse_delimited(data)
    assert len(session.points) == 0
    assert session.notes is not None
    assert "numeric measurement" in session.notes


def test_quoted_csv_with_embedded_commas_in_strings() -> None:
    """csv.reader handles quoted-field embedded commas natively, but
    confirm the parser leaves those fields intact and doesn't mangle
    column alignment."""
    data = (
        b'Obs,Photo,Cond,Ci,PARi,Note\n'
        b'1,15.0,0.25,260,1500,"steady, mid-day reading"\n'
        b'2,14.7,0.24,262,1500,"after, watering"\n'
    )
    session = parse_delimited(data)
    assert len(session.points) == 2
    assert session.points[0].photo_a == pytest.approx(15.0)
    assert session.points[0].raw.get("Note") == "steady, mid-day reading"


def test_mixed_case_and_spaced_headers_still_match_aliases() -> None:
    """Headers like ' photo ' or 'COND' or 'A_Net' — case + whitespace
    must not break the alias lookup."""
    data = (
        b"Obs, photo , COND ,A_Net,CI,PARi\n"
        b"1,15.0,0.25,99.0,260,1500\n"
    )
    session = parse_delimited(data)
    p = session.points[0]
    # First alias-match wins for photo_a (the ' photo ' column).
    assert p.photo_a == pytest.approx(15.0)
    assert p.cond_gsw == pytest.approx(0.25)
    assert p.ci_ppm == pytest.approx(260.0)
    # A_Net (canonical key 'a_net') is also an alias of photo_a; first
    # column wins, A_Net flows into raw.
    assert p.raw.get("A_Net") == pytest.approx(99.0)


def test_to_dict_is_jsonable_and_preserves_structure() -> None:
    session = parse_delimited(_li6400_tsv(num_points=2))
    d = session.to_dict()
    s = json.dumps(d, allow_nan=False)
    assert "NaN" not in s and "Infinity" not in s
    assert d["instrument"] == "li_6400"
    assert len(d["points"]) == 2
