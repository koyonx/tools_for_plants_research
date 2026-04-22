"""Robust LI-COR gas-exchange file parser.

Supports LI-6400 and LI-6800 spreadsheet exports plus generic CSV / TSV
files operators may have hand-edited.  The research group produces
files in every one of these formats depending on which instrument they
grab and whether they transcribe to Excel afterwards, so this parser
has to cope with all of them without operator intervention.

Parsing strategy
----------------
We don't assume a specific sheet name or a specific header row index.
Instead, for each candidate table (an xlsx sheet, or the whole CSV):

1. Read every row into a list-of-lists.
2. Walk the rows looking for the **header row** — the first row that
   contains at least 3 tokens that alias to known LI-COR variables
   (``Photo``, ``A``, ``Ci``, ``Cond``, ``gsw``, …).  Everything
   above is treated as the metadata block.
3. The row immediately below the header may be a **unit row** (LI-6800
   puts ``µmol m-2 s-1`` under each numeric column).  We detect it as
   a row whose cells are mostly strings containing physics units /
   symbols and skip it.
4. Remaining rows are data.  For each column, we normalise the header
   via an alias dictionary, cast numerics where possible, and keep the
   raw cell in the ``raw`` jsonb on the output row so downstream
   analyses can still read instrument-specific fields.

Outputs a `ParsedSession` (normalised metadata + list of
`ParsedPoint`) that the API layer inserts into Supabase.

Deliberately no pandas: the parsing logic is small and explicit, and
keeping pandas out of the runtime image saves ~25 MB.  We use openpyxl
for .xlsx and Python's csv module for text.
"""

from __future__ import annotations

import csv
import io
import math
import re
from dataclasses import asdict, dataclass, field
from datetime import UTC, datetime
from typing import Any

from openpyxl import load_workbook

# Every key is a canonical column name we store in gas_exchange_points.
# Values are the lower-cased token seen in the file header; the parser
# strips spaces, parentheses, units, and LI-6800's multi-line names
# before matching.
#
# The lists are deliberately verbose — LI-COR renames half of their
# columns between firmware revisions (Photo→A, Cond→gsw, PARi→Qin)
# and operators often retype headers when exporting to Excel.  Each
# canonical key maps to one DB column (see migration).
COLUMN_ALIASES: dict[str, tuple[str, ...]] = {
    "photo_a": (
        "photo", "a", "a_net", "anet", "an", "assimilation",
        "photo_avg", "a_avg",
    ),
    "cond_gsw": (
        "cond", "gsw", "gs", "gs_co2", "gh2o", "cond_h2o", "stomatalcond",
    ),
    "ci_ppm": (
        "ci", "c_i", "cint", "ci_ppm", "ci_pa", "intercellularco2",
    ),
    "co2_ref_ppm": (
        "co2r", "ca", "co2_r", "co2ref", "co2reference",
        "reference_co2", "co2_air",
    ),
    "co2_sample_ppm": (
        "co2s", "co2_s", "co2sample", "co2_chamber", "cs",
    ),
    "transpiration": (
        "trmmol", "e", "transpiration", "e_apparent", "e_leaf", "trans",
    ),
    "vpd_kpa": (
        "vpdl", "vpd_l", "vpdleaf", "vpd_leaf", "vpdair", "vpd",
        "vpd_a", "vpd_kpa",
    ),
    "leaf_temp_c": (
        "tleaf", "t_leaf", "tl", "leaftemp", "t_leaf_c",
    ),
    "par_umol": (
        "pari", "qin", "ppfd", "par", "par_in", "q_in", "qamb",
    ),
    "rh_pct": (
        "rhcham", "rh_r", "rh", "rhsample", "rhref", "rh_chamber",
        "relativehumidity",
    ),
    "flow_umol": (
        "flow", "flow_r", "flow_umol", "flow_rate",
    ),
}
# Reverse lookup: token → canonical key, O(1) at parse time.
_TOKEN_TO_KEY: dict[str, str] = {
    token: key for key, tokens in COLUMN_ALIASES.items() for token in tokens
}

# Minimum unique canonical keys a row must contain to be classified as
# the header row.  3 is enough to distinguish a LI-COR header from
# random string rows in the metadata block without missing sparse
# custom exports.
HEADER_MIN_KEYWORDS = 3

# Patterns that strongly suggest a "unit row" directly beneath the
# header (LI-6800 style).  When >50% of a row's non-empty cells match
# one of these, we skip it as a unit row.
_UNIT_PATTERNS: tuple[re.Pattern[str], ...] = (
    re.compile(r"^[µu]mol"),
    re.compile(r"^mol\s"),
    re.compile(r"^mmol"),
    re.compile(r"^kpa", re.I),
    re.compile(r"^pa$", re.I),
    re.compile(r"^°c$|^degc$|^deg\s*c", re.I),
    re.compile(r"^%$|^percent", re.I),
    re.compile(r"^m-2", re.I),
    re.compile(r"^\[.*\]$"),
    re.compile(r"^\(.*\)$"),
)

# Tokens we recognise in a HEADER block to extract the instrument.
# The token must appear in a metadata cell for auto-identification;
# otherwise we fall back to column-alias voting.
_INSTRUMENT_HINTS: dict[str, str] = {
    "li-6400": "li_6400",
    "li6400": "li_6400",
    "li-6800": "li_6800",
    "li6800": "li_6800",
    "licor-6400": "li_6400",
    "licor-6800": "li_6800",
}


@dataclass(frozen=True)
class ParsedPoint:
    obs_index: int
    recorded_at: str | None  # ISO 8601 string; None if the file lacked it
    photo_a: float | None
    cond_gsw: float | None
    ci_ppm: float | None
    co2_ref_ppm: float | None
    co2_sample_ppm: float | None
    transpiration: float | None
    vpd_kpa: float | None
    leaf_temp_c: float | None
    par_umol: float | None
    rh_pct: float | None
    flow_umol: float | None
    raw: dict[str, Any] = field(default_factory=dict)

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


@dataclass(frozen=True)
class ParsedSession:
    instrument: str               # 'li_6400' | 'li_6800' | 'generic_csv' | 'unknown'
    source_format: str            # 'xlsx' | 'csv' | 'tsv' | 'unknown'
    captured_at: str | None
    file_name: str | None
    metadata: dict[str, Any]      # raw header block entries
    notes: str | None
    points: list[ParsedPoint]

    def to_dict(self) -> dict[str, Any]:
        return {
            "instrument": self.instrument,
            "source_format": self.source_format,
            "captured_at": self.captured_at,
            "file_name": self.file_name,
            "metadata": self.metadata,
            "notes": self.notes,
            "points": [p.to_dict() for p in self.points],
        }


class LicorParseError(ValueError):
    """Raised when no header row could be identified."""


def _normalise_token(text: str) -> str:
    """Strip spaces, punctuation, units, and lower-case a header cell so
    we can look it up in `_TOKEN_TO_KEY`.  Preserves underscores so
    operator-typed ``photo_avg`` still matches."""
    s = text.strip().lower()
    # Drop trailing units in parens: "Photo (umol/m2/s)" → "photo"
    s = re.sub(r"\(.*?\)", "", s)
    # Drop trailing units after whitespace: "Photo umol m-2 s-1" → "photo"
    s = re.split(r"\s+(?:[µu]?mol|mol|kpa|ppm|[°d]egc|%)", s, maxsplit=1)[0]
    s = s.strip()
    s = re.sub(r"[^a-z0-9_]", "", s)
    return s


def _coerce_float(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, int | float):
        v = float(value)
        return v if math.isfinite(v) else None
    if isinstance(value, str):
        s = value.strip()
        if not s or s.lower() in {"nan", "na", "n/a", "none", "null", "inf", "-inf", "#n/a", "#value!"}:
            return None
        try:
            v = float(s)
        except ValueError:
            return None
        return v if math.isfinite(v) else None
    return None


def _coerce_datetime(value: Any) -> str | None:
    """Return an ISO-8601 UTC string, or None if the cell isn't parseable.

    LI-COR files carry a mix of ``datetime`` objects (openpyxl) and
    strings (CSV export).  Format priority is **unambiguous-first**:
    we accept ISO-8601 (Y/M/D), then dotted European (D.M.Y), and only
    fall back to slash dates after rejecting the ones whose
    interpretation is ambiguous.

    For slash dates, ``M/D/Y`` and ``D/M/Y`` look identical when both
    fields are <= 12.  Rather than silently mis-parse 50% of those (a
    real research-data hazard — Aug 7 vs Jul 8 changes which growing
    season the measurement belongs to), we return None for ambiguous
    slash dates and let the caller note the issue.  Unambiguous slash
    dates (one field > 12) are still parsed under MM/DD/YYYY since
    that's what LI-COR firmware exports in en-US locales.
    """
    if value is None:
        return None
    if isinstance(value, datetime):
        if value.tzinfo is None:
            value = value.replace(tzinfo=UTC)
        # mypy can't narrow datetime.isoformat() return type without
        # the explicit cast — astimezone keeps type datetime, not
        # datetime[Any].
        return str(value.astimezone(UTC).isoformat())
    if isinstance(value, str):
        s = value.strip()
        if not s:
            return None
        # Try unambiguous formats first: ISO-style (Y first) and
        # dotted European (D.M.Y).  These can never collide.
        for fmt in (
            "%Y-%m-%d %H:%M:%S",
            "%Y-%m-%dT%H:%M:%S",
            "%Y/%m/%d %H:%M:%S",
            "%d.%m.%Y %H:%M:%S",
        ):
            try:
                dt = datetime.strptime(s, fmt).replace(tzinfo=UTC)
                return dt.isoformat()
            except ValueError:
                continue
        # Slash format: only accept when one of the leading fields is
        # > 12, so the M/D vs D/M ambiguity is resolved by data shape.
        slash_match = re.match(
            r"^(\d{1,2})/(\d{1,2})/(\d{4})\s+(\d{1,2}):(\d{2})(?::(\d{2}))?$", s
        )
        if slash_match:
            a, b, y, hh, mm, ss = slash_match.groups()
            ai, bi = int(a), int(b)
            if ai > 12:
                # Must be DD/MM/YYYY (else ai isn't a valid month).
                month, day = bi, ai
            elif bi > 12:
                # Must be MM/DD/YYYY.
                month, day = ai, bi
            else:
                # Ambiguous (both <= 12).  Refuse to guess.
                return None
            try:
                dt = datetime(
                    int(y), month, day, int(hh), int(mm), int(ss or 0), tzinfo=UTC
                )
                return dt.isoformat()
            except ValueError:
                return None
        # Final attempt: pure ISO
        try:
            dt = datetime.fromisoformat(s)
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=UTC)
            return dt.astimezone(UTC).isoformat()
        except ValueError:
            return None
    return None


def _is_unit_row(row: list[Any]) -> bool:
    nonempty = [str(c).strip() for c in row if c is not None and str(c).strip()]
    if not nonempty:
        return False
    hits = 0
    for cell in nonempty:
        for pat in _UNIT_PATTERNS:
            if pat.search(cell):
                hits += 1
                break
    return hits / len(nonempty) > 0.5


def _score_header(
    row: list[Any],
) -> tuple[int, dict[int, str], list[str], list[tuple[int, str]]]:
    """Return (match_count, col_index → canonical_key, raw header
    strings, list of duplicate (col_index, canonical_key)).

    `match_count` is how many canonical keys this row's tokens alias
    to.  When two header columns alias to the same canonical key (e.g.
    a hand-edited file with both ``Photo`` AND ``A`` columns), the
    FIRST occurrence wins for the typed column and the duplicate is
    reported via the fourth tuple element so the caller can keep the
    duplicate's value in the `raw` blob — overwriting silently would
    drop data, which is unsafe given the user explicitly asked for
    robustness across operator-edited variants.
    """
    col_to_key: dict[int, str] = {}
    raw_headers: list[str] = []
    seen_keys: set[str] = set()
    duplicates: list[tuple[int, str]] = []
    for i, cell in enumerate(row):
        if cell is None:
            raw_headers.append("")
            continue
        text = str(cell).strip()
        raw_headers.append(text)
        if not text:
            continue
        token = _normalise_token(text)
        if not token:
            continue
        key = _TOKEN_TO_KEY.get(token)
        if key is None:
            continue
        if key in seen_keys:
            # First column wins; later duplicates flow into raw via
            # the duplicates list so the caller can preserve them.
            duplicates.append((i, key))
            continue
        col_to_key[i] = key
        seen_keys.add(key)
    return len(seen_keys), col_to_key, raw_headers, duplicates


def _find_header_row(
    rows: list[list[Any]],
) -> tuple[int, dict[int, str], list[str], list[tuple[int, str]]]:
    """Pick the first row with the strongest alias match.  Raises if no
    row has >= HEADER_MIN_KEYWORDS known columns."""
    best_idx = -1
    best_score = 0
    best_map: dict[int, str] = {}
    best_headers: list[str] = []
    best_dupes: list[tuple[int, str]] = []
    for idx, row in enumerate(rows):
        score, col_map, raw, dupes = _score_header(row)
        if score > best_score:
            best_score = score
            best_idx = idx
            best_map = col_map
            best_headers = raw
            best_dupes = dupes
            if score >= len(COLUMN_ALIASES):
                break  # perfect match, can't do better
    if best_score < HEADER_MIN_KEYWORDS:
        raise LicorParseError(
            f"Could not identify a LI-COR header row "
            f"(max {best_score} known tokens found, need >= {HEADER_MIN_KEYWORDS}). "
            "Ensure the file is a LI-6400/LI-6800 export or a CSV with "
            "columns like Photo/A, Cond/gsw, Ci, ..."
        )
    return best_idx, best_map, best_headers, best_dupes


def _infer_instrument(
    metadata_rows: list[list[Any]],
    col_to_key: dict[int, str],
    header_raw: list[str],
) -> str:
    """Guess which machine produced the file.

    Priority:
    1. Any metadata-block cell containing a model-name hint
       (``LI-6400`` / ``LI-6800``).
    2. Header-row token presence: ``Photo`` / ``Cond`` are LI-6400
       legacy names; ``A`` / ``gsw`` / ``Qin`` are the LI-6800
       renames.  Both are in the alias map so both parse fine; this
       just labels the source so the UI can show it.
    3. ``generic_csv`` when we recognised columns but can't tell
       which machine; ``unknown`` when even the columns are opaque.
    """
    for row in metadata_rows:
        for cell in row:
            if cell is None:
                continue
            text = str(cell).strip().lower().replace(" ", "")
            for hint, machine in _INSTRUMENT_HINTS.items():
                if hint in text:
                    return machine

    # Column-vote heuristic based on the HEADER row tokens (not the
    # metadata block — LI-6800's metadata usually just says "HEADER"
    # with no model string).  The tokens are already lowercase after
    # normalisation.
    header_tokens = {_normalise_token(h) for h in header_raw if h}
    legacy_tokens = {"photo", "cond", "pari", "trmmol", "vpdl"}  # LI-6400
    modern_tokens = {"a", "gsw", "qin", "vpdleaf"}               # LI-6800
    has_legacy = bool(header_tokens & legacy_tokens)
    has_modern = bool(header_tokens & modern_tokens)
    if has_modern and not has_legacy:
        return "li_6800"
    if has_legacy and not has_modern:
        return "li_6400"
    return "generic_csv" if col_to_key else "unknown"


def _cell_to_json_safe(value: Any) -> Any:
    """Shrink a cell to a JSON-encodable scalar.  openpyxl hands us
    datetime, float, int, str, bool, or None — plus occasional
    formula-result wrappers.  Everything non-finite becomes None."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return _coerce_datetime(value)
    if isinstance(value, bool):
        return value
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return value if math.isfinite(value) else None
    if isinstance(value, str):
        return value
    return str(value)


def _parse_rows(
    rows: list[list[Any]],
    *,
    source_format: str,
    file_name: str | None,
) -> ParsedSession:
    """Shared rows → ParsedSession logic.  `rows` must be a rectangular
    list of lists of cell values (any type); CSV / TSV / xlsx all get
    normalised to this shape before calling."""
    # Strip trailing empty rows so they don't drag through the header
    # search with noisy zero-score matches.
    while rows and all(c is None or str(c).strip() == "" for c in rows[-1]):
        rows.pop()
    if not rows:
        raise LicorParseError("empty file")

    header_idx, col_to_key, raw_headers, duplicate_alias_cols = _find_header_row(rows)
    metadata_rows = rows[:header_idx]
    body_rows = rows[header_idx + 1 :]
    # Column indexes whose canonical key collided with an earlier
    # column.  We know these are alias-of-numeric so we float-coerce
    # them when routing into `raw` (vs. the verbatim string preserved
    # for genuinely unknown columns like operator status flags).
    duplicate_col_idx: set[int] = {idx for idx, _ in duplicate_alias_cols}

    # Skip optional LI-6800 unit row directly under the header.
    if body_rows and _is_unit_row(body_rows[0]):
        body_rows = body_rows[1:]

    instrument = _infer_instrument(metadata_rows, col_to_key, raw_headers)

    # Flatten metadata into a dict{key: value} when possible — the LI-6400
    # HEADER section writes "Key\tValue" pairs; LI-6800 often has longer
    # strings.  Fall back to list-of-lists.
    meta_dict: dict[str, Any] = {"rows": []}
    for row in metadata_rows:
        cells = [_cell_to_json_safe(c) for c in row]
        meta_dict["rows"].append(cells)
        non_empty = [c for c in cells if c not in (None, "")]
        if len(non_empty) == 2 and isinstance(non_empty[0], str):
            meta_dict[str(non_empty[0]).strip()] = non_empty[1]

    # Identify the obs-index / datetime columns up front.
    recorded_col: int | None = None
    for i, h in enumerate(raw_headers):
        token = _normalise_token(h)
        if token in {
            "date",
            "datetime",
            "time",
            "timestamp",
            "hhmmss",
            "fdate",
            "ftime",
        }:
            recorded_col = i
            break
    obs_col: int | None = None
    for i, h in enumerate(raw_headers):
        token = _normalise_token(h)
        if token in {"obs", "obsnum", "obsnumber", "index", "n"}:
            obs_col = i
            break

    points: list[ParsedPoint] = []
    inferred_captured: str | None = None
    for row in body_rows:
        # A body row is considered real if it has at least one
        # numeric-coercible value across the mapped columns.  Otherwise
        # it's a stray blank / "REMARK=" annotation LI-COR injects.
        mapped_values: dict[str, float | None] = {k: None for k in COLUMN_ALIASES}
        any_numeric = False
        for col_idx, key in col_to_key.items():
            if col_idx >= len(row):
                continue
            v = _coerce_float(row[col_idx])
            mapped_values[key] = v
            if v is not None:
                any_numeric = True
        if not any_numeric:
            continue

        recorded_at = None
        if recorded_col is not None and recorded_col < len(row):
            recorded_at = _coerce_datetime(row[recorded_col])
            if recorded_at and inferred_captured is None:
                inferred_captured = recorded_at

        if obs_col is not None and obs_col < len(row):
            oi_raw = row[obs_col]
            oi = int(oi_raw) if isinstance(oi_raw, int | float) else len(points) + 1
        else:
            oi = len(points) + 1

        # Duplicate-alias columns (e.g. an operator-edited file with
        # BOTH `Photo` AND `A`) are intentionally absent from
        # `col_to_key` so the first occurrence wins for the typed
        # column.  Their values still flow into `raw_extras` under
        # the original header so nothing is silently dropped — but we
        # float-coerce them since we KNOW they're alias-of-numeric;
        # genuinely-unknown columns stay verbatim (operator status
        # flags like "ok"/"flagged" must round-trip as strings).
        raw_extras: dict[str, Any] = {}
        for i, header in enumerate(raw_headers):
            if i in col_to_key or i in (recorded_col, obs_col):
                continue
            if not header:
                continue
            if i >= len(row):
                continue
            if i in duplicate_col_idx:
                raw_extras[header] = _coerce_float(row[i])
            else:
                raw_extras[header] = _cell_to_json_safe(row[i])

        points.append(
            ParsedPoint(
                obs_index=oi,
                recorded_at=recorded_at,
                raw=raw_extras,
                **mapped_values,
            )
        )

    captured = inferred_captured
    # Fall back to a metadata block field if nothing in the body had a
    # timestamp.  LI-6400 writes "FileDate".
    for key in ("FileDate", "StartDate", "Date", "Started"):
        if captured:
            break
        val = meta_dict.get(key)
        if val is not None:
            captured = _coerce_datetime(val)

    diagnostic_notes: list[str] = []
    if duplicate_alias_cols:
        # Surface the alias collision so the operator knows which
        # column was treated as ground truth and where to look for
        # the duplicate's values (in `raw`).
        first_idx = {key: idx for idx, key in col_to_key.items()}
        details = ", ".join(
            f"col[{idx}]='{raw_headers[idx]}' lost to col[{first_idx[k]}]='{raw_headers[first_idx[k]]}' for {k}"
            for idx, k in duplicate_alias_cols
        )
        diagnostic_notes.append(
            f"duplicate alias columns detected; first occurrence wins, "
            f"duplicates preserved in `raw`: {details}"
        )
    if len(body_rows) > 0 and len(points) == 0:
        # All body rows existed but every one was non-numeric / NaN.
        # Flag explicitly so the operator doesn't see "0 points" with
        # no clue whether the file was empty or every row was scrubbed.
        diagnostic_notes.append(
            f"file had {len(body_rows)} body rows but none yielded a "
            "numeric measurement value — every row was empty / NaN / "
            "non-coercible.  Check that the source file has cached "
            "formula values (openpyxl cannot evaluate live formulas)."
        )

    return ParsedSession(
        instrument=instrument,
        source_format=source_format,
        captured_at=captured,
        file_name=file_name,
        metadata=meta_dict,
        notes="; ".join(diagnostic_notes) if diagnostic_notes else None,
        points=points,
    )


def parse_xlsx(content: bytes, *, file_name: str | None = None) -> ParsedSession:
    """Parse a LI-6400 / LI-6800 Excel export.

    We scan every sheet and pick the one whose header-row score is
    highest — operators sometimes rename the default sheet, and the
    measurement log can live beside summary pivots.
    """
    wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    best: ParsedSession | None = None
    best_points = -1
    errors: list[str] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows: list[list[Any]] = []
        for row in ws.iter_rows(values_only=True):
            rows.append(list(row))
        try:
            parsed = _parse_rows(rows, source_format="xlsx", file_name=file_name)
        except LicorParseError as e:
            errors.append(f"[{sheet_name}] {e}")
            continue
        if len(parsed.points) > best_points:
            best = parsed
            best_points = len(parsed.points)
    if best is None:
        raise LicorParseError(
            "no LI-COR-compatible sheet found in workbook. "
            + ("; ".join(errors) if errors else "")
        )
    return best


def parse_delimited(
    content: bytes | str,
    *,
    file_name: str | None = None,
    delimiter: str | None = None,
) -> ParsedSession:
    """Parse CSV or TSV text.  Delimiter is sniffed when `delimiter` is
    None — LI-COR exports are tab-delimited by default but operators
    save as CSV."""
    text = content.decode("utf-8", errors="replace") if isinstance(content, bytes) else content
    # Python's Sniffer wants a few lines; give it the first KB.
    probe = text[:2048]
    if delimiter is None:
        try:
            dialect = csv.Sniffer().sniff(probe, delimiters=",\t;")
            delimiter = dialect.delimiter
        except csv.Error:
            # Default to tab — more common in LI-COR raw logs.
            delimiter = "\t" if "\t" in probe else ","
    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    rows = [list(row) for row in reader]
    fmt = "tsv" if delimiter == "\t" else "csv"
    return _parse_rows(rows, source_format=fmt, file_name=file_name)


def parse_file(content: bytes, *, file_name: str | None = None) -> ParsedSession:
    """Dispatch to the right parser based on filename extension + magic.

    xlsx files begin with the zip marker ``PK``; treat anything else as
    text and let `parse_delimited` sniff the separator.  This covers:

    - .xlsx (LI-6800, newer LI-6400 firmware with ExcelExport enabled)
    - .xls (old Excel — not supported; openpyxl cannot read this format,
      raises a clear error so the operator can re-save as .xlsx)
    - .csv / .tsv / .txt (LI-6400 raw log; LI-6800 after File > Export
      as CSV; hand-edited exports).
    """
    name_lower = (file_name or "").lower()
    is_xlsx = name_lower.endswith(".xlsx") or (
        len(content) >= 2 and content[:2] == b"PK"
    )
    if is_xlsx:
        return parse_xlsx(content, file_name=file_name)
    if name_lower.endswith(".xls"):
        raise LicorParseError(
            "legacy .xls (pre-2007) files are not supported. "
            "Re-save the file as .xlsx or export as CSV/TSV."
        )
    return parse_delimited(content, file_name=file_name)
